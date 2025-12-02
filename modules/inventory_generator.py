import os
import glob
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from modules.export_import import build_column_mapping, detect_target_from_headers, import_rows
import re
import unicodedata
from difflib import SequenceMatcher

# Encabezados base para hoja INVENTARIO_GENERAL
HEADERS_INVENTARIO = [
    'ID', 'SERIAL', 'CODIGO_UNIFICADO', 'CODIGO_INDIVIDUAL', 'PLACA', 'ANTERIOR_PLACA',
    'SEDE', 'IP_SEDE', 'UBICACION', 'TECNOLOGIA', 'MARCA', 'MODELO', 'MAC', 'SO',
    'HOSTNAME', 'IP', 'MARCA_MODELO_TELEMEDICINA', 'SERIAL_TELEMEDICINA', 'MOUSE',
    'TECLADO', 'ANTERIOR_ASIGNADO', 'ASIGNADO_A', 'CARGO', 'AREA', 'CONTACTO',
    'FECHA_ASIGNACION', 'CARGADO_NIT', 'ENVIADO_NIT', 'FECHA_ENVIADO', 'GUIA',
    'PROCESADOR', 'ARQUITECTURA_RAM', 'CANTIDAD_RAM', 'TIPO_DISCO', 'ALMACENAMIENTO',
    'PLACA_MONITOR', 'ANTERIOR_PLACA_MONITOR', 'SERIAL_MONITOR', 'MARCA_MONITOR',
    'MODELO_MONITOR', 'ESTADO', 'DISPONIBLE', 'OBSERVACIONES',
    'TIPO_COMPONENTE_ADICIONAL', 'MARCA_MODELO_COMPONENTE_ADICIONAL',
    'SERIAL_COMPONENTE_ADICIONAL', 'MARCA_MODELO_TELEFONO', 'SERIAL_TELEFONO',
    'IMEI_TELEFONO', 'MARCA_MODELO_IMPRESORA', 'IP_IMPRESORA', 'SERIAL_IMPRESORA',
    'PIN_IMPRESORA', 'MARCA_MODELO_CCTV', 'SERIAL_CCTV', 'ENTRADA_OC_COMPRA',
    'FECHA_LLEGADA', 'OC', 'PROVEEDOR', 'MUEBLE_ASIGNADO', 'CREADOR',
    'FECHA_CREACION', 'TRAZABILIDAD', 'ACCIONES', 'FECHA_ULTIMA_MODIFICACION',
    'ARCHIVO_ORIGEN'
]


def _codigo_sede(valor):
    if pd.isna(valor):
        return "XXX"
    codigos = {
        'MEDELLIN': 'MED', 'MEDELLÍN': 'MED', 'CARTAGENA': 'CTG', 'MONTERIA': 'MTR',
        'MONTERÍA': 'MTR', 'PASTO': 'PST', 'BUCARAMANGA': 'BGA', 'IBAGUE': 'IBG',
        'IBAGUÉ': 'IBG', 'CALI': 'CLI', 'BOGOTA': 'BOG', 'BOGOTÁ': 'BOG',
        'CUCUTA': 'CUC', 'CÚCUTA': 'CUC', 'VILLAVICENCIO': 'VVC', 'TUNJA': 'TNJ',
        'PEREIRA': 'PER', 'MANIZALES': 'MNZ'
    }
    txt = str(valor).upper().strip()
    return codigos.get(txt, txt[:3])


def _codigo_tecnologia(valor):
    if pd.isna(valor):
        return "GEN"
    codigos = {
        'PORTATIL': 'PORT', 'PORTÁTIL': 'PORT', 'PC': 'PC', 'COMPUTADOR': 'PC',
        'TODO EN UNO': 'AIO', 'ALL IN ONE': 'AIO', 'MONITOR': 'MON', 'IMPRESORA': 'IMP',
        'TELEFONO': 'TEL', 'TELÉFONO': 'TEL', 'MOUSE': 'MOU', 'TECLADO': 'TEC',
        'CCTV': 'CCTV', 'CAMARA': 'CAM', 'CÁMARA': 'CAM', 'SCANNER': 'SCN',
        'ESCANER': 'SCN', 'ESCÁNER': 'SCN'
    }
    txt = str(valor).upper().strip()
    return codigos.get(txt, txt[:3])


def _estandarizar(df):
    mapa = {
        'Marca': {
            'hewlett-packard': 'HP', 'lenovo': 'Lenovo', 'dell': 'Dell',
            'apple': 'Apple', 'asus': 'Asus', 'acer': 'Acer', 'samsung': 'Samsung',
            'lg': 'LG', 'janus': 'Janus', 'toshiba': 'Toshiba'
        },
        'Tipo de dispositivo': {
            'laptop': 'Portátil', 'portatil': 'Portátil', 'computador de escritorio': 'PC',
            'computador': 'PC', 'todo en uno': 'AIO', 'all in one': 'AIO',
            'impresora': 'Impresora', 'monitor': 'Monitor', 'telefono': 'Teléfono',
            'teléfono': 'Teléfono', 'escaner': 'Scanner', 'escáner': 'Scanner'
        }
    }
    for col, mapeo in mapa.items():
        if col in df.columns:
            df[col] = df[col].astype(str).str.lower().map(mapeo).fillna(df[col])
    return df


def _normalize_header(text: str) -> str:
    """Lowercase, strip BOM/accents/symbols to improve matching."""
    if text is None:
        return ""
    txt = str(text).strip().replace("\ufeff", "")
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    txt = re.sub(r"[^0-9a-zA-Z _-]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip().lower()
    return txt


def _find_header_row(df, max_scan: int = 25):
    """Choose the row with the highest count of text-like entries."""
    best_idx, best_score = 0, -1
    limit = min(len(df), max_scan)
    for i in range(limit):
        row = df.iloc[i]
        text_like = sum(isinstance(x, str) and x.strip() for x in row)
        if text_like > best_score:
            best_score = text_like
            best_idx = i
    return best_idx


def _clean_headers_and_rows(df):
    """
    Detect header row, drop Unnamed/empty columns, and remove leading index column
    if it is purely numeric. Returns cleaned dataframe with normalized headers.
    """
    hdr_idx = _find_header_row(df)
    raw_headers = df.iloc[hdr_idx].tolist()
    headers = [_normalize_header(h) for h in raw_headers]
    df = df.iloc[hdr_idx + 1:].reset_index(drop=True)
    df.columns = headers
    # Remove unnamed/empty headers
    df = df.loc[:, ~df.columns.str.match(r"^unnamed|^$")]
    # Drop first column if numeric-only index
    if df.shape[1] > 0:
        col0 = df.columns[0]
        if df[col0].dropna().apply(lambda x: str(x).isdigit()).all():
            df = df.drop(columns=[col0])
    return df


def _leer_xlsx(upload_folder):
    files = glob.glob(os.path.join(upload_folder, "*.xlsx")) + glob.glob(os.path.join(upload_folder, "*.xls"))
    lista = []
    for path in files:
        try:
            # leer todas las hojas sin asumir encabezado
            sheets = pd.read_excel(path, sheet_name=None, header=None)
        except Exception:
            continue
        for sheet_name, df in sheets.items():
            if df is None or df.empty:
                continue
            df = df.copy()
            # limpieza de encabezados/fila índice
            df = _clean_headers_and_rows(df)
            df["ARCHIVO_ORIGEN"] = os.path.basename(path)
            df["__hoja_origen__"] = sheet_name
            lista.append(df)
    return lista


def _deduplicar(df, columna_serial="Serial o Mac"):
    if columna_serial in df.columns:
        df["__non_null_count__"] = df.count(axis=1)
        df.sort_values(by=[columna_serial, "__non_null_count__"], ascending=[True, False], inplace=True)
        df.drop_duplicates(subset=[columna_serial], keep="first", inplace=True)
        df.drop(columns=["__non_null_count__"], inplace=True)
    return df


def _generar_excel(df_informe, df_sedes, output_path):
    wb = Workbook()
    wb.remove(wb.active)
    c = {'p': "1F4E78", 'wh': "FFFFFF"}
    ws_inv = wb.create_sheet("INVENTARIO_GENERAL")
    # encabezados
    for idx, h in enumerate(HEADERS_INVENTARIO, 1):
        cell = ws_inv.cell(row=1, column=idx)
        cell.value = h
        cell.font = Font(name='Calibri', bold=True, size=9, color=c['wh'])
        cell.fill = PatternFill(start_color=c['p'], end_color=c['p'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    mapeo = {
        'Serial o Mac': 'SERIAL',
        'Sede': 'SEDE',
        'Tipo de dispositivo': 'TECNOLOGIA',
        'Marca': 'MARCA',
        'Modelo': 'MODELO',
        'Hostname': 'HOSTNAME',
        'Mouse': 'MOUSE',
        'Teclado': 'TECLADO',
        'Asignado a': 'ASIGNADO_A',
        'Cargo Asignado': 'CARGO',
        'Area Asignado': 'AREA',
        'Procesador': 'PROCESADOR',
        'Tipo de ram': 'ARQUITECTURA_RAM',
        'Cantidad de ram': 'CANTIDAD_RAM',
        'Tipo de disco': 'TIPO_DISCO',
        'Espacio de disco': 'ALMACENAMIENTO',
        'Estado': 'ESTADO',
        'Observaciones': 'OBSERVACIONES',
        'ARCHIVO_ORIGEN': 'ARCHIVO_ORIGEN',
    }
    sede_ip_map = {}
    if df_sedes is not None and 'Sede' in df_sedes.columns and 'IP_Sede' in df_sedes.columns:
        sede_ip_map = df_sedes.set_index('Sede')['IP_Sede'].to_dict()

    cont_ind, cont_agr = {}, {}
    for ridx, (_, row) in enumerate(df_informe.iterrows(), 2):
        ws_inv.cell(row=ridx, column=1).value = ridx - 1  # ID
        sede_val = row.get('Sede', '')
        tec_val = row.get('Tipo de dispositivo', '')
        usuario_val = row.get('Asignado a', '')

        for col_orig, col_dest in mapeo.items():
            if col_orig in df_informe.columns and pd.notna(row.get(col_orig)):
                ws_inv.cell(row=ridx, column=HEADERS_INVENTARIO.index(col_dest) + 1).value = str(row[col_orig])

        if sede_val in sede_ip_map:
            ws_inv.cell(row=ridx, column=HEADERS_INVENTARIO.index('IP_SEDE') + 1).value = sede_ip_map[sede_val]

        cs = _codigo_sede(sede_val)
        ct = _codigo_tecnologia(tec_val)
        k_ind = f"{cs}-{ct}"
        cont_ind[k_ind] = cont_ind.get(k_ind, 0) + 1
        ci = f"{k_ind}-{cont_ind[k_ind]:03d}"
        ws_inv.cell(row=ridx, column=HEADERS_INVENTARIO.index('CODIGO_INDIVIDUAL') + 1).value = ci
        ws_inv.cell(row=ridx, column=HEADERS_INVENTARIO.index('PLACA') + 1).value = ci

        if usuario_val:
            k_agr = f"{cs}-AGRU"
            cont_agr[k_agr] = cont_agr.get(k_agr, 0) + 1
            ca = f"{k_agr}-{cont_agr[k_agr]:03d}"
            ws_inv.cell(row=ridx, column=HEADERS_INVENTARIO.index('CODIGO_UNIFICADO') + 1).value = ca
            ws_inv.cell(row=ridx, column=HEADERS_INVENTARIO.index('DISPONIBLE') + 1).value = "NO"
        else:
            ws_inv.cell(row=ridx, column=HEADERS_INVENTARIO.index('DISPONIBLE') + 1).value = "SI"

        today = datetime.now().strftime("%d/%m/%Y")
        ws_inv.cell(row=ridx, column=HEADERS_INVENTARIO.index('CREADOR') + 1).value = "Sistema"
        ws_inv.cell(row=ridx, column=HEADERS_INVENTARIO.index('FECHA_CREACION') + 1).value = today
        ws_inv.cell(row=ridx, column=HEADERS_INVENTARIO.index('FECHA_ULTIMA_MODIFICACION') + 1).value = today

    for i in range(1, len(HEADERS_INVENTARIO) + 1):
        ws_inv.column_dimensions[get_column_letter(i)].width = 15
    ws_inv.freeze_panes = "A2"
    ws_inv.auto_filter.ref = ws_inv.dimensions

    # Hoja combinada de materias primas de datos
    ws_comb = wb.create_sheet("COMBINADO")
    for row in dataframe_to_rows(df_informe, index=False, header=True):
        ws_comb.append(list(row))

    wb.save(output_path)


def generar_informe_inventario(upload_folder: str, reports_dir: str):
    try:
        lista_df = _leer_xlsx(upload_folder)
        if not lista_df:
            return None, "No se encontraron archivos .xlsx/.xls válidos."

        # Procesar cada DataFrame individualmente: estandarizar, deduplicar, importar según destino
        cleaned_dfs = []
        stats = []
        for df in lista_df:
            df = _estandarizar(df)
            df = _deduplicar(df, "Serial o Mac")
            headers = list(df.columns)
            mapping, norm_headers = build_column_mapping(headers)
            target = detect_target_from_headers(norm_headers)
            inserted, updated, errors, staged = import_rows(
                df,
                target,
                mapping,
                source=f"{df.get('ARCHIVO_ORIGEN', '')}::{df.get('__hoja_origen__', '')}"
            )
            stats.append({
                "target": target,
                "inserted": inserted,
                "updated": updated,
                "staged": staged,
                "errors": errors[:5],
            })
            cleaned_dfs.append(df)

        df_informe = pd.concat(cleaned_dfs, ignore_index=True, sort=False) if cleaned_dfs else pd.DataFrame()

        # Cargar Sedes si existe en la subida
        df_sedes = None
        path_sedes = os.path.join(upload_folder, "Sedes.xlsx")
        if os.path.exists(path_sedes):
            try:
                df_sedes = pd.read_excel(path_sedes, sheet_name="Sedes")
            except Exception:
                df_sedes = None

        # Generar Excel final
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(reports_dir, f"SISTEMA_DEFINITIVO_INTEGRAL_IPS_{ts}.xlsx")
        os.makedirs(reports_dir, exist_ok=True)
        _generar_excel(df_informe, df_sedes, output_path)

        return output_path, stats
    except Exception as e:
        return None, str(e)
